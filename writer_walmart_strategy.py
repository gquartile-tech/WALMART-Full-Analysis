"""
writer_walmart_strategy.py

Writes the Walmart Account Strategy template from a Walmart Databricks export.

Tab mapping (Walmart vs Amazon equivalent):
  17_Client_Success_Insights_Repo  ←→  38_Client_Success_Insights_Repo
  15_Gong_Call_Insights            ←→  37_Gong_Call_Insights_for_Sales
  14_Project_Dataset_on_SF         ←→  54_Project_Dataset_on_SF
  18_Client_Journey_Insights_Data  ←→  39_Client_Journey_Insights_Data
  09_Campaigns_Grouped_by_Product  ←→  14_Campaign_Performance_by_Adve
  25_Ad_Item_Performance_Report    (Walmart-specific item data)
  01_Advertiser_Name               ←→  01_Advertiser_Name

Key Walmart differences:
  - daily_target_spend__c  (not monthly)
  - Target_ROAS__c is primary KPI (not ACoS)
  - Target_TACoS__c present, Target_ACoS__c often null
  - No catalogue tab — use 25_Ad_Item_Performance_Report for item-level data
  - Product IDs are numeric ItemIds, not ASINs

Safety:
  - _safe_write() handles MergedCell targets
  - _build_cid_row skips MergedCell instances
  - keep_vba=True preserves macros
"""
from __future__ import annotations

import math
import os
import re
import warnings
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment


# ── Utility helpers ───────────────────────────────────────────────────────────

def _safe(val, default: str = "") -> Any:
    """Return default for None/NaN, otherwise val."""
    if val is None:
        return default
    if isinstance(val, float) and math.isnan(val):
        return default
    return val


def _fmt_date(val) -> str:
    if val is None:
        return ""
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    return str(val)


def _safe_write(ws, coord: str, value) -> None:
    """
    Write value to coord, unmerging first if it's a MergedCell.
    """
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        target_range = None
        for merge_range in list(ws.merged_cells.ranges):
            if coord in merge_range:
                target_range = merge_range
                break
        if target_range:
            ws.unmerge_cells(str(target_range))
            min_col_letter = openpyxl.utils.get_column_letter(target_range.min_col)
            top_left = f"{min_col_letter}{target_range.min_row}"
            ws[top_left] = value
            ws[top_left].alignment = Alignment(wrap_text=True, vertical='top')
        else:
            ws[coord] = value
    else:
        ws[coord] = value


# ── Sheet readers ─────────────────────────────────────────────────────────────

def _find_header_row(ws, max_scan: int = 10) -> Optional[int]:
    """Return 1-based row index of the header row (first row with >3 non-null cells)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        if len([c for c in row if c is not None]) > 3:
            return i
    return None


def _tab_to_dict(ws) -> Dict[str, Any]:
    """Single data-row tab → {header: value}."""
    hr = _find_header_row(ws)
    if hr is None:
        return {}
    rows = list(ws.iter_rows(min_row=hr, max_row=hr + 1, values_only=True))
    if len(rows) < 2:
        return {}
    headers, data = rows[0], rows[1]
    return {h: data[i] for i, h in enumerate(headers)
            if h is not None and i < len(data)}


def _tab_to_records(ws) -> List[Dict[str, Any]]:
    """Multi-row tab → list of {header: value} dicts."""
    hr = _find_header_row(ws)
    if hr is None:
        return []
    headers = None
    records = []
    for row in ws.iter_rows(min_row=hr, values_only=True):
        if headers is None:
            headers = list(row)
            continue
        if not any(v is not None for v in row):
            continue
        records.append({
            headers[j]: row[j]
            for j in range(min(len(headers), len(row)))
            if headers[j] is not None
        })
    return records


def _latest_record(records: List[Dict], modstamp_key: str = 'SystemModstamp') -> Dict:
    """Return the record with the most recent SystemModstamp, or first record."""
    if not records:
        return {}
    dated = [(r.get(modstamp_key), r) for r in records if r.get(modstamp_key)]
    if not dated:
        return records[0]
    return max(dated, key=lambda x: x[0])[1]


def _parse_walmart_header(wb) -> Dict[str, str]:
    """
    Walmart exports put identity in 01_Advertiser_Name with a different format
    than the Analysis tabs. Try both.
    """
    # Try 01_Advertiser_Name first (same as Amazon)
    ws01 = next((wb[n] for n in wb.sheetnames if n.startswith('01_Advertiser_Name')), None)
    if ws01:
        rows = list(ws01.iter_rows(min_row=1, max_row=4, values_only=True))
        title      = str(rows[0][0]) if rows else ''
        account_line = str(rows[1][0]) if len(rows) > 1 else ''
        range_line   = str(rows[2][0]) if len(rows) > 2 else ''
        dl_line      = str(rows[3][0]) if len(rows) > 3 else ''

        m_id = re.search(r'Tenant ID:\s*(.*?)\s*\|\s*(?:Account|Advertiser) ID:\s*(\S+)', account_line)
        m_win = re.search(r'(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})', range_line)

        # Account label: strip the " - Advertiser Name" suffix
        hash_name = re.sub(r'\s*-\s*Advertiser[_\s]*Name\s*$', '', title, flags=re.I).strip()

        return {
            'hash_name':   hash_name,
            'tenant_id':   m_id.group(1).strip() if m_id else '',
            'account_id':  m_id.group(2).strip() if m_id else '',
            'date_range':  f"{m_win.group(1)} to {m_win.group(2)}" if m_win else '',
            'downloaded':  dl_line.replace('Downloaded:', '').strip(),
        }

    # Fallback: read from any Analysis sheet header rows
    for candidate in ['Account Mastery_Analysis', 'Framework_Analysis',
                       'Account Health_Analysis', 'Account Strategy _Analysis']:
        if candidate in wb.sheetnames:
            rows = list(wb[candidate].iter_rows(min_row=1, max_row=5, values_only=True))
            title_line   = str(rows[0][0]) if rows else ''
            account_line = str(rows[2][1]) if len(rows) > 2 and rows[2][1] else ''
            range_line   = str(rows[3][1]) if len(rows) > 3 and rows[3][1] else ''
            dl            = rows[4][1]     if len(rows) > 4 else None

            hash_name = re.sub(r'\s*—.*$', '', title_line).strip()
            m_id  = re.search(r'Tenant ID:\s*(.*?)\s*\|\s*Account ID:\s*(\S+)', account_line)
            m_win = re.search(r'(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})', range_line)

            return {
                'hash_name':   hash_name,
                'tenant_id':   m_id.group(1).strip() if m_id else '',
                'account_id':  m_id.group(2).strip() if m_id else '',
                'date_range':  f"{m_win.group(1)} to {m_win.group(2)}" if m_win else '',
                'downloaded':  dl.strftime('%Y-%m-%d %H:%M:%S') if hasattr(dl, 'strftime') else str(dl or ''),
            }

    raise ValueError("No recognizable header sheet found in Walmart export.")


# ── ChildASIN View builder ────────────────────────────────────────────────────

def _build_child_item_view(ws_child, item_records: List[Dict], qt_by_item: Dict) -> None:
    """
    Populate the ChildASIN View tab from Walmart item-level data.

    Walmart equivalent mapping to Amazon ChildASIN View:
      Parent ASIN     → ItemId (Walmart has no parent/child hierarchy — use ItemId)
      ASIN            → ItemId
      Total Sales     → DirectSales + BrandSales + RelatedSales (TotalAdSales as proxy)
      Ad Spend        → AdSpend
      Ad Sales        → TotalAdSales
      Ads Units       → Orders
      Clicks          → Clicks
      ATM_Spend       → sum of ATM spend from qt_by_item
      BA_Spend        → BA spend
      SPT_Spend       → SPT spend
      WATM_Spend      → WATM spend
      SB_Spend        → SB spend
      SBV_Spend       → SBV spend (video)
      SD_Spend        → SD spend
    """
    # Read header row (row 2) to get column positions
    header_to_col: Dict[str, int] = {}
    for cell in ws_child[2]:
        if cell.value and not isinstance(cell, MergedCell):
            header_to_col[cell.value] = cell.column

    # Clear rows 3+ from previous data
    for row in ws_child.iter_rows(min_row=3, max_col=ws_child.max_column):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None

    QT_SPEND_COLS = {
        'ATM_Spend':         {'ATM', 'atm'},
        'BA_Spend':          {'BA', 'ba', 'BRAND', 'brand'},
        'Manual_Q1_Spend':   {'MANUAL', 'manual', 'Manual_Q1'},
        'BAK_Spend':         {'BAK', 'bak'},
        'OP_Spend':          {'OP', 'op'},
        'SPT_Spend':         {'SPT', 'spt'},
        'WATM_Spend':        {'WATM', 'watm'},
        'SB_Spend':          {'SB', 'sb', 'sponsoredbrands'},
        'SBV_Spend':         {'SBV', 'sbv', 'video', 'VIDEO'},
        'SD_Spend':          {'SD', 'sd', 'sponsoreddisplay'},
        'NonQuartile_Spend': {'NONQUARTILE', 'non_quartile', 'nonquartile'},
    }

    def _qt_spend(item_id: str, qt_types: set) -> float:
        total = 0.0
        for qt_type, spend in qt_by_item.get(str(item_id), {}).items():
            if qt_type.upper() in {t.upper() for t in qt_types}:
                total += spend
        return total

    for row_idx, rec in enumerate(item_records, start=3):
        item_id   = _safe(rec.get('ItemId', ''))
        ad_spend  = float(_safe(rec.get('AdSpend', 0), 0) or 0)
        ad_sales  = float(_safe(rec.get('TotalAdSales', 0), 0) or 0)
        orders    = float(_safe(rec.get('Orders', 0), 0) or 0)
        clicks    = float(_safe(rec.get('Clicks', 0), 0) or 0)
        item_name = _safe(rec.get('AdItemName', ''))

        acos = ad_spend / ad_sales if ad_sales > 0 else None
        aov  = ad_sales / orders   if orders   > 0 else None
        tacos = None  # Walmart item-level doesn't have total sales — leave blank

        def _w(header: str, value):
            col = header_to_col.get(header)
            if col:
                cell = ws_child.cell(row=row_idx, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = value

        _w('Parent ASIN',        item_id)
        _w('ASIN',               item_id)
        _w('Total Sales',        ad_sales)   # best proxy available
        _w('Ad Spend',           ad_spend)
        _w('Ad Sales',           ad_sales)
        _w('Ads Units Ordered',  int(orders))
        _w('Clicks',             int(clicks))
        _w('ACoS',               acos)
        _w('AOV',                aov)
        _w('TAG 1',              item_name[:50] if item_name else '')

        # QT spend by campaign type
        for header, qt_types in QT_SPEND_COLS.items():
            _w(header, _qt_spend(item_id, qt_types))

        # Ad Sales % and Organic Sales % — both = 100%/0% since we only have ad data
        _w('Ad Sales (%)',      1.0 if ad_sales > 0 else 0.0)
        _w('Organic Sales (%)', 0.0)

        # Quartile One and Bulk formulas
        q1_col   = header_to_col.get('Quartile One')
        qbulk_col = header_to_col.get('Quartile Bulk')
        o_col = header_to_col.get('ATM_Spend', 0)        # O
        q_col = header_to_col.get('Manual_Q1_Spend', 0)  # Q
        s_col = header_to_col.get('OP_Spend', 0)         # S

        if q1_col:
            o_l = openpyxl.utils.get_column_letter(o_col) if o_col else 'O'
            q_l = openpyxl.utils.get_column_letter(q_col) if q_col else 'Q'
            s_l = openpyxl.utils.get_column_letter(s_col) if s_col else 'S'
            ws_child.cell(row=row_idx, column=q1_col).value = (
                f'=SUM({o_l}{row_idx}+{q_l}{row_idx}+{s_l}{row_idx})'
            )

        ba_col   = header_to_col.get('BA_Spend', 0)
        bak_col  = header_to_col.get('BAK_Spend', 0)
        spt_col  = header_to_col.get('SPT_Spend', 0)
        watm_col = header_to_col.get('WATM_Spend', 0)
        sb_col   = header_to_col.get('SB_Spend', 0)
        sbv_col  = header_to_col.get('SBV_Spend', 0)
        sd_col   = header_to_col.get('SD_Spend', 0)
        imp_col  = header_to_col.get('Imported_Spend', 0)
        nq_col   = header_to_col.get('NonQuartile_Spend', 0)

        if qbulk_col:
            p_l  = openpyxl.utils.get_column_letter(ba_col)   if ba_col   else 'P'
            r_l  = openpyxl.utils.get_column_letter(bak_col)  if bak_col  else 'R'
            t_l  = openpyxl.utils.get_column_letter(spt_col)  if spt_col  else 'T'
            v_l  = openpyxl.utils.get_column_letter(watm_col) if watm_col else 'V'
            w_l  = openpyxl.utils.get_column_letter(sb_col)   if sb_col   else 'W'
            x_l  = openpyxl.utils.get_column_letter(sbv_col)  if sbv_col  else 'X'
            y_l  = openpyxl.utils.get_column_letter(sd_col)   if sd_col   else 'Y'
            z_l  = openpyxl.utils.get_column_letter(imp_col)  if imp_col  else 'Z'
            aa_l = openpyxl.utils.get_column_letter(nq_col)   if nq_col   else 'AA'
            ws_child.cell(row=row_idx, column=qbulk_col).value = (
                f'=SUM({p_l}{row_idx}+{r_l}{row_idx}+{t_l}{row_idx}+'
                f'{v_l}{row_idx}+{w_l}{row_idx}+{x_l}{row_idx}+'
                f'{y_l}{row_idx}+{z_l}{row_idx}+{aa_l}{row_idx})'
            )


# ── Main writer ───────────────────────────────────────────────────────────────

def write_walmart_strategy(export_path: str, template_path: str, output_dir: str) -> str:
    """
    Read a Walmart Databricks export and fill the Strategy template.
    Returns the path of the saved output file.
    """
    # ── Open export ───────────────────────────────────────────────────────────
    pa = openpyxl.load_workbook(export_path, data_only=True, read_only=True)

    # ── Identity ──────────────────────────────────────────────────────────────
    hdr = _parse_walmart_header(pa)
    hash_name  = hdr['hash_name']
    tenant_id  = hdr['tenant_id']
    account_id = hdr['account_id']
    date_range = hdr['date_range']
    downloaded = hdr['downloaded']

    # ── Tab 17: Client Success Insights ──────────────────────────────────────
    ws17 = next((pa[n] for n in pa.sheetnames if n.startswith('17_')), None)
    d17  = _tab_to_dict(ws17) if ws17 else {}

    # ── Tab 14: Project Dataset on SF ─────────────────────────────────────────
    ws14 = next((pa[n] for n in pa.sheetnames if n.startswith('14_')), None)
    proj_records = _tab_to_records(ws14) if ws14 else []
    # Filter to Walmart channel row for this account
    proj_row = {}
    if proj_records:
        wm_rows = [r for r in proj_records
                   if str(r.get('Channel__c', '')).strip().lower() == 'walmart'
                   and str(r.get('Advertiser_ID__c', '')).strip() == str(account_id).strip()]
        proj_row = wm_rows[0] if wm_rows else _latest_record(proj_records)

    # ── Tab 15: Gong ──────────────────────────────────────────────────────────
    ws15 = next((pa[n] for n in pa.sheetnames if n.startswith('15_')), None)
    gong_records = _tab_to_records(ws15) if ws15 else []
    # Sort by call end date desc, take most recent
    gong_records_sorted = sorted(
        gong_records,
        key=lambda r: r.get('Gong__Call_End__c') or datetime.min,
        reverse=True,
    )
    gong = gong_records_sorted[0] if gong_records_sorted else {}

    # ── Tab 18: Client Journey ────────────────────────────────────────────────
    ws18 = next((pa[n] for n in pa.sheetnames if n.startswith('18_')), None)
    cjm  = _tab_to_dict(ws18) if ws18 else {}

    # ── Tab 25: Ad Item Performance ───────────────────────────────────────────
    ws25 = next((pa[n] for n in pa.sheetnames if n.startswith('25_')), None)
    item_records = _tab_to_records(ws25) if ws25 else []

    # ── Tab 09: Campaigns Grouped by Product + QT Type ────────────────────────
    # Build a dict: {item_id: {qt_type: spend}} for ChildASIN View
    ws09 = next((pa[n] for n in pa.sheetnames if n.startswith('09_')), None)
    qt_records = _tab_to_records(ws09) if ws09 else []
    qt_by_item: Dict[str, Dict[str, float]] = {}
    for rec in qt_records:
        item_id  = str(_safe(rec.get('ProductId', ''), ''))
        qt_type  = str(_safe(rec.get('QuartileCampaignType', ''), ''))
        spend    = float(_safe(rec.get('Spend', 0), 0) or 0)
        if item_id and qt_type and item_id.lower() != 'multiple':
            qt_by_item.setdefault(item_id, {})[qt_type] = (
                qt_by_item.get(item_id, {}).get(qt_type, 0) + spend
            )

    # ── Tab 16: Stripe ────────────────────────────────────────────────────────
    ws16 = next((pa[n] for n in pa.sheetnames if n.startswith('16_')), None)
    stripe_records = _tab_to_records(ws16) if ws16 else []
    # Most recent paid charge
    stripe = next(
        (r for r in sorted(stripe_records,
                            key=lambda r: r.get('PaymentDate') or datetime.min,
                            reverse=True)
         if str(r.get('PaymentStatus', '')).lower() == 'paid'),
        {}
    )

    pa.close()

    # ── Customer age from CSP ─────────────────────────────────────────────────
    customer_age_months = _safe(d17.get('Customer_Age_Months__c'))

    # ── Load template ─────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(template_path, keep_vba=True)

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 1 — Questionaire Survey - AMZ  (same sheet name kept in Walmart template)
    # ════════════════════════════════════════════════════════════════════════════
    ws1 = wb['Questionaire Survey - AMZ']

    def w(coord, value):
        _safe_write(ws1, coord, value)

    # ── Client overview block ─────────────────────────────────────────────────
    # Row 6
    w('C6', hash_name)                                          # Member ID / Account
    w('E6', account_id)                                         # Channel/Advertiser ID
    w('I6', _safe(d17.get('CSM__c')))                          # Last Modified By / CSM

    # Row 7
    w('E7', account_id)                                         # Primary Advertiser Hash
    w('I7', _safe(proj_row.get('daily_target_spend__c')))      # Projected MRR / Daily Spend

    # Row 8
    w('C8', hash_name)                                          # Account Name
    w('I8', f"{customer_age_months} months" if customer_age_months else '')  # Project Age

    # Row 9
    w('C9', customer_age_months)                                # Customer Age (months)
    w('E9', _safe(d17.get('Repeat_Purchase_Behavior__c')))     # Repeat Purchase Behavior
    w('I9', _safe(d17.get('CSM_Churn_Risk__c')))               # CSM Churn Risk

    # Row 10
    w('C10', _safe(d17.get('Commodity_Products_or_Branded_Products__c')))  # Commodity/Branded
    w('E10', _safe(d17.get('Sales_Concentration__c')))          # Sales Concentration
    w('I10', _safe(d17.get('Director_Churn_Risk__c')))          # Director Churn Risk

    # Row 11
    w('C11', _safe(d17.get('CSM__c')))                         # Current CSM
    w('E11', _safe(d17.get('CSM_Tenure__c')))                  # CSM Tenure
    w('I11', _safe(d17.get('Account_Risk_Score__c')))          # Account Risk Score

    # Row 12
    w('C12', '')                                                # Director (not in Walmart export)
    w('E12', _safe(cjm.get('Active_Products__c')))             # Active Products

    # Row 13
    w('E13', _safe(d17.get('Customer_Feedback__c')))           # Overall Satisfaction

    # ── Strategy block ────────────────────────────────────────────────────────
    # Row 15
    w('C15', _safe(d17.get('Current_Challenges__c')))          # Current Challenges
    w('E15', _safe(d17.get('Primary_Objective__c')))           # Primary Objective
    w('I15', _safe(d17.get('ACOS_Constraint__c')))             # ACoS Constraint

    # Row 16
    w('C16', _safe(d17.get('Primary_Objective_Additional_Context__c')))  # Objective Context
    w('E16', _safe(d17.get('Primary_Spend_KPI__c')))           # Primary Spend KPI
    w('I16', _safe(d17.get('Customer_Acquisition_Cost_Target__c')))      # CAC Target

    # Row 17
    w('C17', _safe(d17.get('Top_Priority__c')))                # Top Priority
    w('I17', _safe(d17.get('TACOS_Constraint__c')))            # TACoS Constraint

    # Row 18
    w('C18', _safe(d17.get('Second_Priority__c')))             # 2nd Priority
    w('F18', _safe(proj_row.get('CS_Notes__c')))               # CS Notes (from project)
    w('I18', _safe(proj_row.get('daily_target_spend__c')))     # Daily Target Spend

    # Row 19
    w('C19', _safe(d17.get('Biggest_Expansion_Opportunity__c')))          # Expansion Opportunity
    w('E19', _safe(d17.get('Near_Term_3_Month_Considerations__c')))        # Near-Term Considerations
    w('I19', _safe(proj_row.get('Target_ROAS__c')))            # Target ROAS

    # ── Client Journey Map (CJM) stages ──────────────────────────────────────
    stage_rows = {1: (24, 25), 2: (27, 28), 3: (30, 31), 4: (33, 34)}
    for s, (r_adopt, r_date) in stage_rows.items():
        adopt   = _safe(cjm.get(f'AdoptionOrUpsellS{s}__c'))
        strategy= _safe(cjm.get(f'StrategyS{s}__c'))
        status  = _safe(cjm.get(f'StatusS{s}__c'))
        exec_dt = cjm.get(f'ExecutionDateS{s}__c')

        w(f'C{r_adopt}', adopt)
        w(f'G{r_adopt}', strategy)
        w(f'I{r_adopt}', status)
        w(f'C{r_date}',  _fmt_date(exec_dt))

    # ── Gong section ──────────────────────────────────────────────────────────
    w('C41', _safe(gong.get('Gong__Call_Brief__c')))
    w('C42', _safe(gong.get('Gong__Call_Key_Points__c')))
    w('C43', _safe(gong.get('Gong__Call_Highlights_Next_Steps__c')))

    # Optional: write call title and date as context above brief
    call_title = _safe(gong.get('Gong__Title__c'))
    call_end   = gong.get('Gong__Call_End__c')
    call_link  = _safe(gong.get('Gong__View_call__c'))
    if call_title:
        try:
            w('C39', f"{call_title} — {_fmt_date(call_end)}")
        except Exception:
            pass
    if call_link:
        try:
            w('E39', call_link)
        except Exception:
            pass

    # ── Stripe / billing context ───────────────────────────────────────────────
    if stripe:
        try:
            w('I13', f"${float(_safe(stripe.get('Amount', 0), 0)):,.0f} — {_safe(stripe.get('Description'))}")
        except Exception:
            pass

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 2 — Account Strategy _Analysis header
    # ════════════════════════════════════════════════════════════════════════════
    ws2 = wb['Account Strategy _Analysis']
    _safe_write(ws2, 'A1', f"{hash_name} — Account Strategy Analysis")
    _safe_write(ws2, 'B3',
                f"Account: {hash_name} | Tenant ID: {tenant_id} | Account ID: {account_id}")
    _safe_write(ws2, 'B4', date_range)
    _safe_write(ws2, 'B5', downloaded)

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 3 — ChildASIN View  (item-level performance data)
    # ════════════════════════════════════════════════════════════════════════════
    if item_records:
        ws3 = wb['ChildASIN View']
        _build_child_item_view(ws3, item_records, qt_by_item)

    # ── Save ──────────────────────────────────────────────────────────────────
    safe_name = re.sub(r'[<>:"/\\|?*]', '-', hash_name)
    filename  = f"{safe_name} — WM Strategy Analysis {date_range}.xlsm"
    out_path  = os.path.join(output_dir, filename)

    wb.save(out_path)
    try:
        wb.close()
    except Exception:
        pass

    print(f"[writer_walmart_strategy] Saved: {out_path}")
    return out_path


# ── Standalone CLI ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Usage: python writer_walmart_strategy.py <export.xlsx> <template.xlsm> [output_dir]")
        sys.exit(1)
    write_walmart_strategy(
        export_path=sys.argv[1],
        template_path=sys.argv[2],
        output_dir=sys.argv[3] if len(sys.argv) > 3 else ".",
    )
