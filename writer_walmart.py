"""
writer_walmart.py
Writes STATUS, What We Saw, Why It Matters to the Reference tab of each
Walmart CoE template.

Key safety features:
  - _safe_write()  : detects MergedCell targets, unmerges before writing
  - cid_to_row scan: skips MergedCell instances (no .value attribute crash)
  - keep_vba=True  : preserves macros in .xlsm templates
"""
from __future__ import annotations

from typing import Dict

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell

from reader_databricks_walmart import clean_text, pct_str


# ── Shared helpers ────────────────────────────────────────────────────────────

def _safe_write(ws, cell_ref: str, value) -> None:
    """
    Write `value` to `cell_ref`, safely handling merged cells.
    If the target is a MergedCell, find its merge range, unmerge it first,
    then write to the top-left cell of that range.
    """
    cell = ws[cell_ref]

    if isinstance(cell, MergedCell):
        # Find the merge range that owns this cell
        target_range = None
        for merge_range in list(ws.merged_cells.ranges):
            if cell_ref in merge_range:
                target_range = merge_range
                break

        if target_range:
            ws.unmerge_cells(str(target_range))
            # Write to the top-left of the former merge range
            top_left = f"{target_range.min_column_letter if hasattr(target_range, 'min_column_letter') else cell_ref[0]}{target_range.min_row}"
            ws[top_left] = value
            ws[top_left].alignment = Alignment(wrap_text=True, vertical='top')
        else:
            # Fallback: just assign — may still fail but at least we tried
            ws[cell_ref] = value
    else:
        ws[cell_ref] = value


def _build_cid_row_map(ws_ref) -> Dict[str, int]:
    """
    Scan column B of the reference sheet and return {control_id: row_number}.
    Skips MergedCell instances to avoid AttributeError on .value.
    """
    cid_to_row: Dict[str, int] = {}
    for row in range(2, ws_ref.max_row + 1):
        cell = ws_ref.cell(row=row, column=2)   # column B
        if isinstance(cell, MergedCell):
            continue
        cid = clean_text(cell.value).upper()
        if cid and cid.startswith('WM-'):
            cid_to_row[cid] = row
    return cid_to_row


def _write_results_to_ref(ws_ref, results: Dict) -> None:
    """
    Write STATUS (col D), What We Saw (col H), Why It Matters (col I)
    to the Reference tab for every control in `results`.
    """
    cid_to_row = _build_cid_row_map(ws_ref)

    for cid, res in results.items():
        cid_upper = cid.upper()
        if cid_upper not in cid_to_row:
            print(f"[writer] WARNING: {cid} not found in reference tab — skipping.")
            continue
        rr = cid_to_row[cid_upper]

        _safe_write(ws_ref, f'D{rr}', res.status)
        _safe_write(ws_ref, f'H{rr}', res.what)
        _safe_write(ws_ref, f'I{rr}', res.why)

        # Ensure wrap on the two text columns (D is short — no wrap needed)
        for col_letter in ('H', 'I'):
            cell = ws_ref[f'{col_letter}{rr}']
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(wrap_text=True, vertical='top')


# ── Mastery writer ────────────────────────────────────────────────────────────

def write_mastery_output(template_path: str, output_path: str,
                         results: Dict, ctx) -> None:
    wb      = load_workbook(template_path, keep_vba=True)
    ws_main = wb['Account Mastery_Analysis']
    ws_ref  = wb['Account Mastery_Reference']

    # Header block
    _safe_write(ws_main, 'A1', f"{ctx.hash_name} — Account Mastery Analysis")
    _safe_write(ws_main, 'B3',
                f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.account_id}")
    if ctx.window_start and ctx.window_end and ctx.window_days:
        _safe_write(ws_main, 'B4',
                    f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)")
    if ctx.downloaded:
        _safe_write(ws_main, 'B5', ctx.downloaded)
        cell = ws_main['B5']
        if not isinstance(cell, MergedCell):
            cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    # Reference tab
    _write_results_to_ref(ws_ref, results)

    wb.save(output_path)
    try:
        wb.close()
    except Exception:
        pass


# ── Framework writer ──────────────────────────────────────────────────────────

def write_framework_output(template_path: str, output_path: str,
                            results: Dict, ctx) -> None:
    wb      = load_workbook(template_path, keep_vba=True)
    ws_main = wb['Framework_Analysis']
    ws_ref  = wb['Framework_Reference']

    _safe_write(ws_main, 'A1', f"{ctx.hash_name} — Framework Analysis")
    _safe_write(ws_main, 'B3',
                f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.account_id}")
    if ctx.window_start and ctx.window_end and ctx.window_days:
        _safe_write(ws_main, 'B4',
                    f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)")
    if ctx.downloaded:
        _safe_write(ws_main, 'B5', ctx.downloaded)
        cell = ws_main['B5']
        if not isinstance(cell, MergedCell):
            cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    _write_results_to_ref(ws_ref, results)

    wb.save(output_path)
    try:
        wb.close()
    except Exception:
        pass


# ── Health writer ─────────────────────────────────────────────────────────────

def write_health_output(template_path: str, output_path: str,
                        results: Dict, ctx) -> None:
    wb      = load_workbook(template_path, keep_vba=True)
    ws_main = wb['Account Health_Analysis']
    ws_ref  = wb['Account Health_Reference']

    _safe_write(ws_main, 'A1', f"{ctx.hash_name} — Account Health Analysis")
    _safe_write(ws_main, 'B3',
                f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.account_id}")
    if ctx.window_start and ctx.window_end and ctx.window_days:
        _safe_write(ws_main, 'B4',
                    f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)")
    if ctx.downloaded:
        _safe_write(ws_main, 'B5', ctx.downloaded)
        cell = ws_main['B5']
        if not isinstance(cell, MergedCell):
            cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    # Health-specific constraint cells
    if ctx.proj_acos_target is not None:
        _safe_write(ws_main, 'D10', pct_str(ctx.proj_acos_target))
    if ctx.proj_tacos_target is not None:
        _safe_write(ws_main, 'D11', pct_str(ctx.proj_tacos_target))

    _write_results_to_ref(ws_ref, results)

    wb.save(output_path)
    try:
        wb.close()
    except Exception:
        pass
