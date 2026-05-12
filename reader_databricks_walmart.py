from __future__ import annotations

import math
import re
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Header row is row index 5 (0-based) by default in Walmart Databricks exports.
# Add per-sheet overrides here if needed.
# ---------------------------------------------------------------------------
SHEET_HEADER_ROW: dict[str, int] = {}
DEFAULT_HEADER_ROW = 5


# ---------------------------------------------------------------------------
# Context object — shared by all three Walmart pillar agents
# ---------------------------------------------------------------------------
@dataclass
class WalmartContext:
    path: str

    # ── Identity ──────────────────────────────────────────────────────────────
    hash_name:   str = ''
    tenant_id:   str = ''
    account_id:  str = ''
    window_start: object = None
    window_end:   object = None
    downloaded:   object = None
    window_days:  Optional[int] = None
    ref_date:     object = None

    # ── CS / Project fields ───────────────────────────────────────────────────
    cs_objective:    str = ''   # advertising objective narrative
    cs_challenges:   str = ''   # account challenges narrative
    cs_seasonality:  str = ''   # seasonality narrative
    cs_constraints:  str = ''   # operational constraints
    cs_customizations: str = '' # customization notes
    journey_stage:   str = ''   # client journey stage
    proj_acos_target:  object = None   # target ACoS (float or None)
    proj_roas_target:  object = None   # target ROAS
    proj_tacos_target: object = None   # target TACoS
    proj_primary_kpi:  str = ''
    proj_notes:        str = ''

    # ── KPIs (02_Date_Range_KPIs) ─────────────────────────────────────────────
    acos:          Optional[float] = None
    prev_acos:     Optional[float] = None
    tacos:         Optional[float] = None
    prev_tacos:    Optional[float] = None
    ad_sales:      Optional[float] = None
    prev_ad_sales: Optional[float] = None
    total_sales:   Optional[float] = None
    ad_spend:      Optional[float] = None
    prev_ad_spend: Optional[float] = None
    ctr:           Optional[float] = None
    prev_ctr:      Optional[float] = None
    cvr:           Optional[float] = None
    prev_cvr:      Optional[float] = None
    cpc:           Optional[float] = None
    prev_cpc:      Optional[float] = None
    roas:          Optional[float] = None
    prev_roas:     Optional[float] = None

    # ── DataFrames ────────────────────────────────────────────────────────────
    df_campaigns:         Optional[pd.DataFrame] = None   # 06_Campaign_Report
    df_campaigns_grouped: Optional[pd.DataFrame] = None   # 07_Campaigns_Grouped_by_Campaig
    df_campaign_meta:     Optional[pd.DataFrame] = None   # 34_Campaign_Metadata
    df_ad_items:          Optional[pd.DataFrame] = None   # 33_Ad_Item_Metadata
    df_ad_groups:         Optional[pd.DataFrame] = None   # 31_Ad_Group_Metadata
    df_product_catalog:   Optional[pd.DataFrame] = None   # 22_Product_Catalog
    df_keywords:          Optional[pd.DataFrame] = None   # 20_Keyword_Performance_Report
    df_item_keywords:     Optional[pd.DataFrame] = None   # 23_Item_Keyword_Performance_Rep
    df_placements:        Optional[pd.DataFrame] = None   # 28_Campaign_Placement_Settings
    df_sd_line_items:     Optional[pd.DataFrame] = None   # 35_SD_Line_Item_Report
    df_advertiser:        Optional[pd.DataFrame] = None   # 12_Advertiser_Settings
    df_product_acos:      Optional[pd.DataFrame] = None   # 13_Product_Level_ACoS
    df_yearly:            Optional[pd.DataFrame] = None   # 03_Yearly_KPIs
    df_monthly_yoy:       Optional[pd.DataFrame] = None   # 05_Monthly_Sales_YoY_Comparison
    df_l24m:              Optional[pd.DataFrame] = None   # 04_L24M_Monthly_Performance_Sum
    df_marketplace_orders: Optional[pd.DataFrame] = None  # 29_Marketplace_Order_Lines
    df_gong:              Optional[pd.DataFrame] = None   # 15_Gong_Call_Insights
    df_qt_campaigns:      Optional[pd.DataFrame] = None   # 08_Campaigns_Grouped_by_QT_Camp

    # ── Derived / computed ────────────────────────────────────────────────────
    gong_gap_days:   Optional[int] = None
    gong_last_call:  object = None


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def clean_text(v) -> str:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ''
    return str(v).replace('&#39;', "'").strip()


def to_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not pd.isna(v):
        return float(v)
    s = clean_text(v)
    if not s or s.lower() in {'nan', 'none', 'null', '-'}:
        return None
    s = s.replace('$', '').replace(',', '').strip()
    if s.endswith('%'):
        try:
            return float(s[:-1]) / 100.0
        except Exception:
            return None
    m = re.match(r'^([0-9]*\.?[0-9]+)k$', s, re.I)
    if m:
        return float(m.group(1)) * 1000.0
    try:
        return float(s)
    except Exception:
        return None


def norm_pct(v) -> Optional[float]:
    x = to_float(v)
    if x is None:
        return None
    return x if x <= 1.0 else x / 100.0


def pct_str(v: Optional[float], decimals: int = 1) -> str:
    if v is None:
        return 'Not documented'
    return f'{v * 100:.{decimals}f}%'


def money_str(v: Optional[float]) -> str:
    if v is None:
        return 'Not documented'
    return f'${v:,.0f}'


def trim(s: str, n: int = 260) -> str:
    s = re.sub(r'\s+', ' ', s or '').strip()
    return s if len(s) <= n else s[:n - 1].rstrip() + '…'


def _find_sheet(wb, prefix: str) -> Optional[str]:
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return name
    return None


def _get_ws(wb, prefix: str):
    name = _find_sheet(wb, prefix)
    return wb[name] if name else None


def _get_df(wb, sheet_prefix: str) -> Optional[pd.DataFrame]:
    sheet = _find_sheet(wb, sheet_prefix)
    if sheet is None:
        return None
    try:
        header_row = SHEET_HEADER_ROW.get(sheet_prefix, DEFAULT_HEADER_ROW)
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= header_row:
            return None
        headers = [str(c) if c is not None else f'Unnamed_{i}' for i, c in enumerate(rows[header_row])]
        data = rows[header_row + 1:]
        df = pd.DataFrame(data, columns=headers)
        df = df.dropna(how='all')
        return df if not df.empty else None
    except Exception:
        return None


def _find_col(df: pd.DataFrame, *candidates: str) -> Optional[str]:
    norm = {re.sub(r'[\s_]', '', str(c)).lower(): c for c in df.columns}
    for cand in candidates:
        key = re.sub(r'[\s_]', '', cand).lower()
        if key in norm:
            return norm[key]
    return None


def _col_val(df: pd.DataFrame, col: Optional[str], row_idx: int = 0):
    if col is None or df is None or df.empty or col not in df.columns:
        return None
    try:
        v = df.iloc[row_idx][col]
        return None if pd.isna(v) else v
    except Exception:
        return None


def _latest_row(df: pd.DataFrame) -> pd.Series:
    for col in df.columns:
        if 'modstamp' in str(col).lower() or 'systemmod' in str(col).lower():
            try:
                df2 = df.copy()
                df2['_ts'] = pd.to_datetime(df2[col], errors='coerce')
                valid = df2.dropna(subset=['_ts'])
                if not valid.empty:
                    return valid.loc[valid['_ts'].idxmax()].drop(labels=['_ts'])
            except Exception:
                pass
    return df.iloc[0]


# ---------------------------------------------------------------------------
# Header parser — reads account identity from the Analysis tab header rows
# (rows 3–5, 1-based) which is the Walmart export format.
# ---------------------------------------------------------------------------

def _parse_walmart_header(wb) -> dict:
    """
    Walmart Databricks exports use 01_Advertiser_Name — same structure as Amazon:
      Row 1: "{AccountName} - Advertiser_Name"
      Row 2: "Account: {name} | Tenant ID: {tid} | Advertiser ID: {aid}"
      Row 3: "Date Range: YYYY-MM-DD to YYYY-MM-DD"
      Row 4: "Downloaded: YYYY-MM-DD HH:MM:SS"
    """
    ws = next((wb[n] for n in wb.sheetnames if n.startswith('01_')), None)
    if ws is None:
        raise ValueError(
            'Sheet 01_Advertiser_Name not found in export. '
            f'Available sheets: {wb.sheetnames[:8]}'
        )

    rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))

    # Row 1 — account name (strip " - Advertiser_Name" suffix)
    title_line = clean_text(rows[0][0]) if rows else ''
    hash_name  = re.sub(r'\s*-\s*Advertiser[_\s]*Name\s*$', '', title_line, flags=re.I).strip()

    # Row 2 — Tenant ID + Advertiser ID
    account_line = clean_text(rows[1][0]) if len(rows) > 1 else ''
    m_id = re.search(
        r'Tenant ID:\s*([\w\-]+)\s*\|\s*(?:Account|Advertiser) ID:\s*(\S+)',
        account_line
    )
    tenant_id  = m_id.group(1).strip() if m_id else ''
    account_id = m_id.group(2).strip() if m_id else ''

    # Row 3 — Date range
    range_line = clean_text(rows[2][0]) if len(rows) > 2 else ''
    m_win = re.search(r'(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})', range_line)
    start = datetime.strptime(m_win.group(1), '%Y-%m-%d').date() if m_win else None
    end   = datetime.strptime(m_win.group(2), '%Y-%m-%d').date() if m_win else None

    # Row 4 — Download datetime
    dl_line    = clean_text(rows[3][0]) if len(rows) > 3 else ''
    dl_str     = re.sub(r'^Downloaded:\s*', '', dl_line, flags=re.I).strip()
    downloaded = None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            downloaded = datetime.strptime(dl_str, fmt)
            break
        except ValueError:
            pass

    # Fallback: row 6/7 have structured data (TenantId, AdvertiserId cols)
    if not tenant_id or not account_id:
        for row in rows[5:]:
            vals = [clean_text(v) for v in row if v is not None]
            if len(vals) >= 2 and re.match(r'[0-9a-f\-]{30,}', vals[0]):
                tenant_id  = tenant_id  or vals[0]
                account_id = account_id or vals[1]
                break

    return {
        'hash_name':    hash_name,
        'tenant_id':    tenant_id,
        'account_id':   account_id,
        'window_start': start,
        'window_end':   end,
        'downloaded':   downloaded,
        'window_days':  (end - start).days + 1 if start and end else None,
        'ref_date':     downloaded.date() if downloaded else end,
    }


# ---------------------------------------------------------------------------
# CS / Project field extraction helpers
# ---------------------------------------------------------------------------

def _read_cs_row(wb) -> Optional[pd.Series]:
    ws = _get_ws(wb, '17_Client_Success_Insights_Repo')
    if ws is None:
        return None
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= DEFAULT_HEADER_ROW:
        return None
    headers = [str(c) if c is not None else f'Unnamed_{i}' for i, c in enumerate(rows[DEFAULT_HEADER_ROW])]
    data = rows[DEFAULT_HEADER_ROW + 1:]
    df = pd.DataFrame(data, columns=headers)
    df = df.dropna(how='all')
    if df.empty:
        return None
    return _latest_row(df)


def _read_proj_row(wb, account_id: str) -> Optional[pd.Series]:
    ws = _get_ws(wb, '14_Project_Dataset_on_SF')
    if ws is None:
        return None
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= DEFAULT_HEADER_ROW:
        return None
    headers = [str(c) if c is not None else f'Unnamed_{i}' for i, c in enumerate(rows[DEFAULT_HEADER_ROW])]
    data = rows[DEFAULT_HEADER_ROW + 1:]
    df = pd.DataFrame(data, columns=headers)
    df = df.dropna(how='all')
    if df.empty:
        return None
    adv_col = _find_col(df, 'Advertiser_ID_c', 'Advertiser_ID', 'AdvertiserID')
    if adv_col and account_id:
        matched = df[df[adv_col].astype(str).str.strip() == str(account_id).strip()]
        df = matched if not matched.empty else df
    return _latest_row(df)


def _read_kpis(df: Optional[pd.DataFrame]) -> dict:
    """Extract KPI fields from 02_Date_Range_KPIs row 0."""
    out: dict = {}
    if df is None or df.empty:
        return out
    row = df.iloc[0]

    def _get(*names):
        for n in names:
            col = _find_col(df, n)
            if col:
                return to_float(row[col])
        return None

    out['acos']          = norm_pct(_get('ACoS', 'ACOS', 'acos'))
    out['prev_acos']     = norm_pct(_get('Prev_ACoS', 'PrevACoS', 'previous_acos'))
    out['tacos']         = norm_pct(_get('TACoS', 'TACOS', 'tacos'))
    out['prev_tacos']    = norm_pct(_get('Prev_TACoS', 'PrevTACoS'))
    out['ad_sales']      = _get('AdSales', 'Ad_Sales', 'adsales')
    out['prev_ad_sales'] = _get('Prev_AdSales', 'PrevAdSales')
    out['total_sales']   = _get('TotalSales', 'Total_Sales', 'totalsales')
    out['ad_spend']      = _get('AdSpend', 'Ad_Spend', 'adspend', 'Spend')
    out['prev_ad_spend'] = _get('Prev_AdSpend', 'PrevAdSpend')
    out['ctr']           = norm_pct(_get('CTR', 'ctr'))
    out['prev_ctr']      = norm_pct(_get('Prev_CTR', 'PrevCTR'))
    out['cvr']           = norm_pct(_get('CR', 'CVR', 'ConversionRate', 'cvr'))
    out['prev_cvr']      = norm_pct(_get('Prev_CR', 'PrevCR', 'Prev_CVR'))
    out['cpc']           = _get('CPC', 'cpc')
    out['prev_cpc']      = _get('Prev_CPC', 'PrevCPC')

    if out['acos'] and out['acos'] > 0:
        out['roas'] = 1.0 / out['acos']
    if out['prev_acos'] and out['prev_acos'] > 0:
        out['prev_roas'] = 1.0 / out['prev_acos']

    return out


# ---------------------------------------------------------------------------
# Main loader
# ---------------------------------------------------------------------------

def load_walmart_context(path: str) -> WalmartContext:
    wb = load_workbook(path, data_only=True, read_only=True)

    try:
        h = _parse_walmart_header(wb)
        ctx = WalmartContext(path=path, **h)

        # ── DataFrames ────────────────────────────────────────────────────────
        ctx.df_campaigns         = _get_df(wb, '06_Campaign_Report')
        ctx.df_campaigns_grouped = _get_df(wb, '07_Campaigns_Grouped_by_Campaig')
        ctx.df_campaign_meta     = _get_df(wb, '34_Campaign_Metadata')
        ctx.df_ad_items          = _get_df(wb, '33_Ad_Item_Metadata')
        ctx.df_ad_groups         = _get_df(wb, '31_Ad_Group_Metadata')
        ctx.df_product_catalog   = _get_df(wb, '22_Product_Catalog')
        ctx.df_keywords          = _get_df(wb, '20_Keyword_Performance_Report')
        ctx.df_item_keywords     = _get_df(wb, '23_Item_Keyword_Performance_Rep')
        ctx.df_placements        = _get_df(wb, '28_Campaign_Placement_Settings')
        ctx.df_sd_line_items     = _get_df(wb, '35_SD_Line_Item_Report')
        ctx.df_advertiser        = _get_df(wb, '12_Advertiser_Settings')
        ctx.df_product_acos      = _get_df(wb, '13_Product_Level_ACoS')
        ctx.df_yearly            = _get_df(wb, '03_Yearly_KPIs')
        ctx.df_monthly_yoy       = _get_df(wb, '05_Monthly_Sales_YoY_Comparison')
        ctx.df_l24m              = _get_df(wb, '04_L24M_Monthly_Performance_Sum')
        ctx.df_marketplace_orders= _get_df(wb, '29_Marketplace_Order_Lines')
        ctx.df_gong              = _get_df(wb, '15_Gong_Call_Insights')
        ctx.df_qt_campaigns      = _get_df(wb, '08_Campaigns_Grouped_by_QT_Camp')

        # ── KPIs ──────────────────────────────────────────────────────────────
        df02 = _get_df(wb, '02_Date_Range_KPIs')
        kpis = _read_kpis(df02)
        for k, v in kpis.items():
            setattr(ctx, k, v)

        # ── CS row ────────────────────────────────────────────────────────────
        cs_row = _read_cs_row(wb)
        if cs_row is not None:
            def _cs(df_cs, *names):
                for n in names:
                    col = _find_col(pd.DataFrame([cs_row]), n)
                    if col:
                        return clean_text(cs_row[col])
                # fallback: positional via col letter
                return ''
            # Read by column name candidates
            for attr, candidates in [
                ('cs_objective',      ['Objective', 'Advertising_Objective', 'AdvertisingObjective']),
                ('cs_challenges',     ['Challenges', 'AccountChallenges', 'Account_Challenges']),
                ('cs_seasonality',    ['Seasonality', 'SeasonalityNotes']),
                ('cs_constraints',    ['Constraints', 'OperationalConstraints', 'Operational_Constraints']),
                ('cs_customizations', ['Customizations', 'Notes', 'CSNotes']),
            ]:
                for cand in candidates:
                    key = re.sub(r'[\s_]', '', cand).lower()
                    matched_col = next(
                        (c for c in cs_row.index if re.sub(r'[\s_]', '', str(c)).lower() == key),
                        None
                    )
                    if matched_col:
                        setattr(ctx, attr, clean_text(cs_row[matched_col]))
                        break

        # ── Journey stage ─────────────────────────────────────────────────────
        ws_journey = _get_ws(wb, '18_Client_Journey_Insights_Data')
        if ws_journey is not None:
            df_j = _get_df(wb, '18_Client_Journey_Insights_Data')
            if df_j is not None and not df_j.empty:
                stage_col = _find_col(df_j, 'Stage', 'JourneyStage', 'Journey_Stage', 'ClientStage')
                if stage_col:
                    ctx.journey_stage = clean_text(df_j.iloc[0][stage_col])

        # ── Project dataset ───────────────────────────────────────────────────
        proj_row = _read_proj_row(wb, ctx.account_id)
        if proj_row is not None:
            for attr, candidates in [
                ('proj_acos_target',  ['ACoS_Target', 'ACoSTarget', 'TargetACoS', 'Target_ACoS']),
                ('proj_roas_target',  ['ROAS_Target', 'ROASTarget', 'TargetROAS', 'Target_ROAS']),
                ('proj_tacos_target', ['TACoS_Target', 'TACOSTarget', 'TargetTACoS', 'Target_TACoS']),
                ('proj_primary_kpi',  ['Primary_KPI', 'PrimaryKPI', 'KPI']),
                ('proj_notes',        ['Notes', 'CSNotes', 'CS_Notes']),
            ]:
                for cand in candidates:
                    key = re.sub(r'[\s_]', '', cand).lower()
                    matched_col = next(
                        (c for c in proj_row.index if re.sub(r'[\s_]', '', str(c)).lower() == key),
                        None
                    )
                    if matched_col:
                        val = proj_row[matched_col]
                        if attr in ('proj_acos_target', 'proj_roas_target', 'proj_tacos_target'):
                            setattr(ctx, attr, norm_pct(val))
                        else:
                            setattr(ctx, attr, clean_text(val))
                        break

        # ── Gong call gap ─────────────────────────────────────────────────────
        if ctx.df_gong is not None and not ctx.df_gong.empty:
            date_col = _find_col(ctx.df_gong, 'CallDate', 'Call_Date', 'Gong__Call_End__c', 'Date')
            if date_col:
                dates = pd.to_datetime(ctx.df_gong[date_col], errors='coerce').dropna().sort_values()
                if len(dates) >= 1:
                    ctx.gong_last_call = dates.iloc[-1]
                    if len(dates) >= 2:
                        ctx.gong_gap_days = int((dates.iloc[-1] - dates.iloc[-2]).days)

    finally:
        try:
            wb.close()
        except Exception:
            pass

    return ctx
